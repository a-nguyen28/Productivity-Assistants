import openpyxl
from openpyxl.styles import Alignment  # Add this import at the top
from openpyxl.styles import numbers  # for predefined number formats


from datetime import datetime, timedelta
import datetime as dtmod
import time
import os



# === CONFIG ===
EXCEL_FILE = "Deep Focused Work Tracker.xlsx"   # Your Excel file
CURRENT_WEEK = "Week 4"            

CATEGORIES = [
    "Classes",
    "Personal",
    "Coursework",
    "Altium",
    "Machine Learning",
    "Academic Clubs",
    "Combat Clubs"
]
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

# === TIMER STATE ===
start_time = None
active_category = None

# === EXCEL HELPERS ===

def get_existing_timedelta(cell_value):
    # Direct numeric values (Excel stores time as fraction of a day)
    if isinstance(cell_value, (int, float)):
        return timedelta(seconds=float(cell_value) * 24 * 3600)

    # datetime.time objects (from Excel cells) -> convert to timedelta
    if isinstance(cell_value, dtmod.time):
        return timedelta(hours=cell_value.hour, minutes=cell_value.minute, seconds=cell_value.second)

    # Try numeric strings (e.g. '02.907000') which may be present
    try:
        s = str(cell_value).strip()
        # empty or zero-like
        if s == "" or s in ("0", "0.0", "0:00:00"):
            return timedelta()
        # If string is a plain float/decimal, interpret heuristically
        f = float(s)
    except Exception:
        # fallback to parsing time-like strings
        return parse_time_string(str(cell_value))
    else:
        # Heuristic: if value is less than 1, treat as Excel day fraction
        if abs(f) < 1:
            return timedelta(seconds=f * 24 * 3600)
        # Otherwise treat as hours (e.g. 2.907 -> 2.907 hours)
        return timedelta(seconds=f * 3600)
    
def get_sheet():
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"Excel file '{EXCEL_FILE}' not found.")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if CURRENT_WEEK not in wb.sheetnames:
        raise ValueError(f"Sheet '{CURRENT_WEEK}' not found in workbook.")
    return wb, wb[CURRENT_WEEK]

def parse_time_string(time_str):
    """Convert '2:56:35' to timedelta."""
    if not time_str or time_str.strip() == "" or time_str == "0:00:00":
        return timedelta()

    s = time_str.strip()
    # If it's a plain decimal string like '02.907000' or '2.907', parse as hours/days
    try:
        f = float(s)
    except Exception:
        pass
    else:
        # treat <1 as Excel day fraction, otherwise as hours
        if abs(f) < 1:
            return timedelta(seconds=f * 24 * 3600)
        return timedelta(seconds=f * 3600)

    # Parse colon-separated H:M:S or M:S
    parts = [p.strip() for p in s.split(":") if p.strip() != ""]
    try:
        if len(parts) == 3:
            h = int(float(parts[0]))
            m = int(float(parts[1]))
            sec = int(float(parts[2]))
        elif len(parts) == 2:
            h = 0
            m = int(float(parts[0]))
            sec = int(float(parts[1]))
        elif len(parts) == 1:
            # single value that isn't plain float (fallback)
            sec = int(float(parts[0]))
            h = 0
            m = 0
        else:
            raise ValueError(f"Unrecognized time format: '{time_str}'")
    except Exception as e:
        raise ValueError(f"Couldn't parse time string '{time_str}': {e}")

    return timedelta(hours=h, minutes=m, seconds=sec)

def format_timedelta(td):
    """Convert timedelta to 'H:MM:SS' string."""
    total_seconds = int(td.total_seconds())
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h}:{m:02}:{s:02}"



def update_excel(category, day, elapsed):
    wb, ws = get_sheet()

    cat_row = CATEGORIES.index(category) + 2
    day_col = DAYS.index(day) + 2

    cell = ws.cell(row=cat_row, column=day_col)

    # Safely get existing time as timedelta
    td_existing = get_existing_timedelta(cell.value)
    existing_days = td_existing.total_seconds() / (24*3600)

    # Convert elapsed to Excel days
    elapsed_days = elapsed.total_seconds() / (24*3600)

    # Add and store back
    new_value = existing_days + elapsed_days
    cell.value = new_value
    cell.number_format = '[h]:mm:ss'
    cell.alignment = Alignment(horizontal='right')

    wb.save(EXCEL_FILE)
    print(f"✅ Added {format_timedelta(elapsed)} to '{category}' on {day}.")
    # Total seconds in a day
    total_seconds = new_value * 24 * 60 * 60
    
    # Compute hours, minutes, seconds
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    seconds = int(total_seconds % 60)
    

    print(f"{category} is now at {hours}:{minutes}:{seconds}")



def fix_existing_durations():
    wb, ws = get_sheet()

    for row in range(2, 2 + len(CATEGORIES)):  # Categories are on rows 2 to ...
        for col in range(2, 2 + len(DAYS)):    # Days are in columns B to H (2 to 8)
            cell = ws.cell(row=row, column=col)
            value = cell.value

            if value is None:
                continue

            # If value is already a float (Excel time), just apply formatting
            if isinstance(value, (int, float)):
                cell.number_format = '[h]:mm:ss'
                cell.alignment = Alignment(horizontal='right')
                continue

            try:
                # Try parsing string time (like "1:23:45") to timedelta
                td = get_existing_timedelta(cell.value)
                cell.value = td.total_seconds() / (24*3600)
                cell.number_format = '[h]:mm:ss'
                cell.alignment = Alignment(horizontal='right')

            except Exception as e:
                print(f"⚠️ Could not process cell at R{row}C{col} with value '{value}': {e}")

    wb.save(EXCEL_FILE)
    print(f"✅ Fixed formatting of existing duration cells in '{CURRENT_WEEK}'.")



# === TIMER LOGIC ===
def start(category):
    global start_time, active_category
    if active_category:
        print(f"⚠️ Timer already running for '{active_category}'. Stop it first.")
        return
    if category not in CATEGORIES:
        print(f"❌ Invalid category '{category}'. Valid options: {CATEGORIES}")
        return
    start_time = time.time()
    active_category = category
    print(f"⏱️ Started tracking '{category}'")

def stop():
    global start_time, active_category
    if not active_category:
        print("⚠️ No timer running.")
        return
    elapsed = timedelta(seconds=time.time() - start_time)
    print(elapsed)
    today = DAYS[datetime.today().weekday()]
    print(f"🛑 Stopped '{active_category}' — Duration: {format_timedelta(elapsed)}")
    update_excel(active_category, today, elapsed)
    start_time = None
    active_category = None

def status():
    if active_category:
        elapsed = timedelta(seconds=time.time() - start_time)
        print(f"⏳ Tracking '{active_category}' — Elapsed: {format_timedelta(elapsed)}")
    else:
        print("No timer currently running.")

# === INTERACTIVE CLI ===
def main():
    print(f"\n📘 Excel Time Tracker — {CURRENT_WEEK}")
    print("Type 'start <category>', 'stop', 'status', 'exit'")
    print(f"Categories: {', '.join(CATEGORIES)}\n")
    while True:
        try:
            cmd = input("> ").strip().lower()
            if not cmd:
                continue

            if cmd.startswith("start "):
                category = cmd.split(" ", 1)[1].title()
                start(category)
            elif cmd == "stop":
                stop()
            elif cmd == "status":
                status()
            elif cmd == "exit":
                print("👋 Goodbye!")
                break
            else:
                print("Unknown command. Try: start <category>, stop, status, exit")

        except KeyboardInterrupt:
            print("\nInterrupted. Type 'exit' to quit safely.")
        except Exception as e:
            print(f"❌ Error: {e}")

if __name__ == "__main__":
    main()